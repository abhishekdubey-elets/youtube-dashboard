"""Excel (.xlsx) export — a local alternative to Google Sheets.

Builds an in-memory workbook with the same columns as the Google Sheet export
(reusing ``HEADER`` / ``build_row``) so the two exports stay identical.
"""
from __future__ import annotations

import io
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from app.services.google_sheets import HEADER, build_row

# Excel hard limit per cell.
_CELL_LIMIT = 32_767
# Reasonable column widths matching HEADER order (15 columns).
_WIDTHS = [40, 14, 44, 22, 24, 22, 14, 10, 70, 70, 32, 12, 14, 22, 12]


def _clean(value: Any) -> str:
    s = "" if value is None else str(value)
    s = ILLEGAL_CHARACTERS_RE.sub("", s)
    return s[:_CELL_LIMIT]


def build_workbook(videos: Iterable[Any]) -> bytes:
    """Return the bytes of an .xlsx workbook for the given videos."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Transcripts"

    ws.append(HEADER)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for video in videos:
        ws.append([_clean(c) for c in build_row(video)])

    for i, width in enumerate(_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
